#!/usr/bin/env python3
# pylint: disable=global-variable-undefined, line-too-long, invalid-name, abstract-class-instantiated

"""
The goal of this script is to automate creation of monthly report
regarding issues reported in Jira. All results are
exported into Excel file.

The script is only relevant to Team1 and Team2. To use it
for other teams, some constraints need to be modified.

Release date      : 11-03-2020
Version           : v1
Author            : Stilian Stoilov
Email             : stenlytu@gmail.com
Tested on         : macOS Catalina, CentOS-7 with Python 3
PyLint score      : 9.61
"""

import configparser
import datetime
import os
import re
import sys
import openpyxl
import pandas as pd
import urllib3
from jira import JIRA, JIRAError
from openpyxl.styles import Alignment, Font

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


def authenticate():
    """
        This function authenticate to Jira using credentials
        stored in credentials.ini file.
    """
    # Open credentials.ini and read Jira credentials. Easier to ask for forgiveness than permission style.
    try:
        with open("credentials.ini") as config_file:
            config = configparser.RawConfigParser()
            config.read_file(config_file)
    except FileNotFoundError as error:
        print(f"File credentials.ini is missing!\n{error}")
        return False
    except PermissionError as error:
        print(f"File credentials.ini cannot be open for reading!\n{error}")
        return False

    user = config["jirauth"]["User"]
    password = config["jirauth"]["Password"]

    # Authenticate to JIRA without checking Certificate.
    global jira
    try:
        jira = JIRA(basic_auth=(user, password), options={'server':'https://jira-link.com', 'verify': False})
        return True
    except JIRAError as error:
        if error.status_code == 401:
            print(f"Login to JIRA failed. Check your username and password.\n{error}")
        return False


def ngpbugs(file, issue_resolutions):
    """
        This function creates Pivot table for Bugs from the first sheet of given file.
        The Pivot table is saved to temporary file. The content of the file is copied to
        the second sheet of the original file and tmp file is removed.

        :param file: The name of the file
        :type file: str
        :param issue_resolutions: Containing unique list of issue's resolution.
        :type issue_resolutions: list
    """
    # Not included resolutions: Rejected, Incomplete

    # print("Issue resolutions: \n{}".format(issue_resolutions))
    all_resolution = ["Open", "In Progress", "Autor Action", "External Action", "Cannot Reproduce", "Consulting", "Decommitted", "Duplicate", "Won't Fix", "Fixed", "Done", "Not resolved", "Won't Do", "Not Applicable"]
    # List comprehension to filter the different element from two lists.
    resolution_for_deletion = [x for x in all_resolution if x not in issue_resolutions]
    resolution_legend = [
        [
            "Unresolved",
            [
                "Open",
                "In Progress",
                "Autor Action",
                "External Action",
                "Not resolved",
                "Won't Do",
            ],
        ],
        ["Resolved", ["Fixed", "Done"]],
        [
            "Consulted",
            [
                "Cannot Reproduce",
                "Consulting",
                "Decommitted",
                "Duplicate",
                "Won't Fix",
                "Not Applicable",
            ],
        ],
    ]

    # print("Resolution for deletion: \n{}".format(resolution_for_deletion))
    for _ in range(3):
        for column in range(3):
            for item in resolution_legend[column][1]:
                if item in resolution_for_deletion:
                    resolution_legend[column][1].remove(item)
    # print("Resolution Legend after: \nResolved - {}, Unresolved - {}, Consulted - {} ".format(resolution_legend[1][1], resolution_legend[0][1], resolution_legend[2][1]))

    # Read the content of the file and create Pivot table.
    try:
        info = pd.read_excel(file, 0)
        p_table = pd.pivot_table(info, index='Component/s', values='Key', columns='Resolution', margins=True, aggfunc='count', fill_value=0)
    except ValueError:
        print(f"\033[91mJira has returned nothing from your query:\033[0m \nQuery: {query}")
        sys.exit(1)

    # Group columns. Issues without a single component are not calculated!
    for index in range(3):
        if len(resolution_legend[index][1]) > 0:
            res_list = []
            for item in resolution_legend[index][1]:
                res_list.append("p_table[\"{}\"]".format(item))
            col = " + ".join(res_list)
            p_table[resolution_legend[index][0]] = eval(col)
            p_table.drop(columns=resolution_legend[index][1], inplace=True)
        else:
            p_table[resolution_legend[index][0]] = 0

    # Revert column order to reach desired output.
    p_table = p_table[p_table.columns.tolist()[::-1]]

    # Start working on rows for Team1.
    p_table = p_table.reset_index().replace({'Component/s': r'.*(?=.*?PostgreSQL)(?=.*?Team1)(?!.*(Team2|RabbitMQ|MongoDB|Redis)).*'}, {'Component/s': 'PostgreSQL'}, regex=True).groupby('Component/s', sort=False).sum()
    p_table = p_table.reset_index().replace({'Component/s': r'.*(?=.*?Redis)(?=.*?(Team1))(?!.*(Team2|PostgreSQL|MongoDB|RabbitMQ)).*'}, {'Component/s': 'Redis'}, regex=True).groupby('Component/s', sort=False).sum()
    p_table = p_table.reset_index().replace({'Component/s': r'.*(?=.*?RabbitMQ)(?=.*?(Team1))(?!.*(Team2|PostgreSQL|MongoDB|Redis)).*'}, {'Component/s': 'RabbitMQ'}, regex=True).groupby('Component/s', sort=False).sum()
    p_table = p_table.reset_index().replace({'Component/s': r'.*(?=.*?MongoDB)(?=.*?(Team1))(?!.*(Team2|PostgreSQL|RabbitMQ|Redis)).*'}, {'Component/s': 'MongoDB'}, regex=True).groupby('Component/s', sort=False).sum()
    p_table = p_table.reset_index().replace({'Component/s': r'.*(?=.*?(Service-Fabrik))(?=.*?(Team1))(?!.*Team2).*'}, {'Component/s': 'Team1'}, regex=True).groupby('Component/s', sort=False).sum()

    # Team2
    p_table = p_table.reset_index().replace({'Component/s': r'.*(?=.*?(Authentication))(?=.*?(Team2))(?!.*Team1).*'}, {'Component/s': 'Authentication'}, regex=True).groupby('Component/s', sort=False).sum()
    p_table = p_table.reset_index().replace({'Component/s': r'.*(?=.*?(Infrastructure))(?=.*?(Team2))(?!.*Team1).*'}, {'Component/s': 'Infrastructure'}, regex=True).groupby('Component/s', sort=False).sum()
    p_table = p_table.reset_index().replace({'Component/s': r'.*(?=.*?(Runtime))(?=.*?(Team2))(?!.*Team1).*'}, {'Component/s': 'Runtime'}, regex=True).groupby('Component/s', sort=False).sum()
    p_table = p_table.reset_index().replace({'Component/s': r'.*(?=.*?(BOSH))(?=.*?(Team2))(?!.*Team1).*'}, {'Component/s': 'BOSH'}, regex=True).groupby('Component/s', sort=False).sum()

    # All rows which have one of the following items are merged with Team2.
    se_core_list = ['CF-Onboarding Service', 'Commercial Infrastructure', 'CF-Platform Monitoring', 'Cloud Engineering', 'CF-Routing', 'CF-Ops', 'CF-AWS', 'CF-Cloud Cockpit']
    for comp in se_core_list:
        regex = '.*(?=.*?' + comp + ')(?=.*?Team2)(?!.*Team1).*'
        p_table = p_table.reset_index().replace({'Component/s': regex}, {'Component/s': 'Team2'}, regex=True).groupby('Component/s', sort=False).sum()

    sorted_row_legend = ['PostgreSQL', 'Redis', 'RabbitMQ', 'MongoDB', 'Service-Fabrik', 'Team1', 'Authentication', 'Infrastructure', 'BOSH', 'Runtime', 'Team2', 'Control Plane', 'All']
    row_list = p_table.index.tolist()
    rows_for_deletion = [x for x in sorted_row_legend if x not in row_list]
    # print("sorted_row_legend: \n{}\nrows_for_deletion: {}".format(sorted_row_legend, rows_for_deletion))

    # Remove missing rows from sorted_row_legend.
    for row_index in sorted_row_legend:
        # print(row_index)Infrastructure!!!
        if row_index in rows_for_deletion:
            sorted_row_legend.remove(row_index)
    # print("Sorted row legend after: \n{} ".format(sorted_row_legend))

    # Re-arrange rows to match desired output.
    p_table = p_table.reindex(sorted_row_legend)

    # Writes the Pivot table to sheet called Pivot.
    with pd.ExcelWriter(path=file, engine="openpyxl", mode="a") as writer:
        p_table.to_excel(writer, sheet_name="Pivot")
        print("\033[32mPivot table generated successfully.\033[0m")

    # Draw and add Bar Chart.
    p_table.drop(columns="All", index="All").plot(stacked=True, kind="bar", figsize=(10, 8), title="Bugs", width=0.6).figure.savefig("figure.png", bbox_inches="tight")
    wb = openpyxl.load_workbook(file)
    ws = wb["Pivot"]
    my_png = openpyxl.drawing.image.Image("figure.png")
    ws.add_image(my_png, "K5")
    wb.save(file)
    os.remove(str(os.getcwd()) + "/figure.png")


def outages():
    """
        TBD
    """
    pass


def main():
    """
        Fetch information from JIRA and writes it into excel file.
    """
    # Check if authentication is successfull.
    if authenticate() is False:
        sys.exit(1)

    # Open configuration.ini and read General settings. EAFP style.
    try:
        with open("configuration.ini") as config_file:
            config = configparser.ConfigParser()
            config.read_file(config_file)
    except FileNotFoundError as error:
        print(f"File configuration.ini is missing!\n{error}")
        sys.exit(1)
    except PermissionError as error:
        print(f"File configuration.ini cannot be open for reading!\n{error}")
        sys.exit(1)
    report_list = list(config.items('Queries'))

    # Iterate over Query list.
    for number, key in enumerate(report_list):
        print(f"\033[93m{str(number + 1)}: {str(key[0])}\033[0m")

    while True:
        select_report = input("\033[36mPlease select report between 1 and {}:\033[0m ".format(len(report_list)))
        try:
            if 1 <= int(select_report) <= len(report_list):
                break
        except ValueError:
            continue

    report_name = report_list[int(select_report) - 1][0]
    global query
    query = report_list[int(select_report) - 1][1]
    today = datetime.date.today()
    filename = "{}_report_{}.xlsx".format(report_name.replace(" ", "_"), today)
    headers = ["Key", "Summary", "Component/s", "Labels", "Resolution"]

    # Start working with the excel file.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jira"
    ws.append(headers)

    # Collect information from Jira filtered by specific JQL. Only the first 500 matches.
    resolution_list = []
    try:
        for issue in jira.search_issues(query, maxResults=500):
            labels = ", ".join(issue.fields.labels)
            component_regex = re.compile(r"name=\'(.*?)\'")
            compon_list = re.findall(component_regex, str(issue.fields.components))
            components = ", ".join(compon_list)
            try:
                resolution = issue.fields.resolution.name
            except AttributeError:
                resolution = "Not resolved"
            finally:
                resolution_list.append(resolution)

            issue_params = [issue.key, issue.fields.summary, components, labels, resolution]
            ws.append(issue_params)
    except JIRAError as error:
        print(f"\033[91mSeems like your query is not correct:\033[0m \nQuery: {query} \n{error}")
        sys.exit(1)

    # Get list of unique issue's resolution.
    resolution_list = list(set(resolution_list))

    # Save the Excel file.
    wb.save(filename)

    # Create the Pivot table.
    ngpbugs(filename, resolution_list)

    # Formating rows and columns.
    bold_font = Font(size=14, bold=True)
    center_aligned_text = Alignment(horizontal="center")
    wb_format = openpyxl.load_workbook(filename)
    for sheet in wb_format.sheetnames:
        ws = wb_format[sheet]
        if sheet == "Jira":
            ws.sheet_properties.tabColor = "1072BA"
            for cell in ws["1:1"]:
                cell.font = bold_font
                cell.alignment = center_aligned_text
        for column_cells in ws.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells) + 2
            ws.column_dimensions[column_cells[0].column_letter].width = length
    wb_format.save(filename)


if __name__ == "__main__":
    main()
